#!/usr/bin/env python2

import openpyxl
import os
import re
import sys

from tempfile import mkstemp
from shutil import move
from os import fdopen, remove


# #


def replace(file_path, pattern, subst):
    # Create temp file
    fh, abs_path = mkstemp()
    with fdopen(fh, 'w') as new_file:
        with open(file_path) as old_file:
            for line in old_file:
			    match = re.search(pattern, line)
			    if match:
			        new_file.write(line.replace(line, subst))
			    else:
			        new_file.write(line.replace('', ''))
    # Remove original file
    remove(file_path)
    # Move new file
    move(abs_path, file_path)


# #

# Get the current directory
source_dir = os.path.dirname(os.path.abspath(__file__))

book = openpyxl.load_workbook(source_dir + '/Random_platform_and_truck_positions.xlsx')

sheet = book.active

# r = 1+1
if len(sys.argv) == 2:
    row = int(sys.argv[1])
else:
    row = 0

r = row+1

if row == 0:  # default
    plat_x = 100
    plat_y = 0
    plat_z = 1
    tr_x = 0
    tr_y= -50

else:
    plat_x = sheet.cell(row=r, column=2).value
    plat_y = sheet.cell(row=r, column=3).value
    plat_z = sheet.cell(row=r, column=4).value
    tr_x = sheet.cell(row=r, column=7).value
    tr_y = sheet.cell(row=r, column=8).value

print 'platform_x =', plat_x
print 'platform_y =', plat_y
print 'platform_z =', plat_z

print 'truck_x =', tr_x
print 'truck_y =', tr_y
# 

replace('{}/{}/dronecourse-eval'.format(source_dir, '../posix-configs/SITL/init/lpe'), 
        'platform',  
        'dronecourse platform {0} {1} {2}\n'.format(plat_x, plat_y, plat_z))
replace('{}/{}/dronecourse-eval.world'.format(source_dir, '../Tools/sitl_gazebo/worlds'), 
        'truck location',  
        '      <pose frame=''>{0} {1} 0 0 0 0</pose> <!-- truck location -->\n'.format(tr_x, tr_y))